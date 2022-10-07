from openpyxl import load_workbook
import sys
from alive_progress import alive_bar


#Loading the workbook
fileName = sys.argv[1] #Get argument 1 as the filename
print("Loading the workbook...")
wb = load_workbook(filename = fileName+'.xlsx')
sheet_ranges = wb['Sheet1']
row_count = sheet_ranges.max_row
print("Total number of enteries found = "+str(row_count))


#Checking if the threshold is given
if(len(sys.argv)>2):
    threshold = int(sys.argv[2]) #Get argument 2 as threshold value
else:
    threshold = 0


#Creating second sheet to save the ranges
ws2 = wb.create_sheet(title="Ranges")
countOfWs2 = 2
countInRange = 1
ws2['A1']="From"
ws2['B1']="To"
ws2['C1']="Valid data count within the range"


#Finding the ranges
print("Finding ranges...")
ws2['A2']=sheet_ranges['A1'].value
with alive_bar(row_count) as bar:
    bar()
    for i in range(1,row_count):
        if((int(sheet_ranges['A'+str(i)].value)+1 == int(sheet_ranges['A'+str(i+1)].value))
           or ((int(sheet_ranges['A'+str(i)].value)+threshold >= int(sheet_ranges['A'+str(i+1)].value)))
           or ((int(sheet_ranges['A'+str(i)].value)+threshold+1 == int(sheet_ranges['A'+str(i+1)].value)))):
            countInRange = countInRange + 1
            ws2['C'+str(countOfWs2)] = countInRange
        else:
            ws2['B'+str(countOfWs2)] = sheet_ranges['A'+str(i)].value
            countOfWs2 = countOfWs2+1
            ws2['A'+str(countOfWs2)] = sheet_ranges['A'+str(i+1)].value
            countInRange = 1
            ws2['C'+str(countOfWs2)] = countInRange
        bar()
    bar()
ws2['B'+str(countOfWs2)] = sheet_ranges['A'+str(i+1)].value


#Saving the file
print("Ranges found, creating and saving as another file...")
wb.save(filename = "Ranges of "+fileName+".xlsx")
print("The ranges file has "+str(countOfWs2)+" ranges.")
print("All done! :D File saved as \"Ranges of "+fileName+".xlsx\"\nYou will find the ranges in the sheet \"Ranges\"")
